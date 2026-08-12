import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

with open('scratch/analysis_results.json', 'r', encoding='utf-8') as f:
    results = json.load(f)

table_data = results['table']

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Danh sách Trường chuẩn hóa"

# Enable grid lines
ws.views.sheetView[0].showGridLines = True

headers = [
    "STT", 
    "Tên Trường (Excel)", 
    "Nguồn File", 
    "Địa Chỉ Chuẩn Hóa Thực Tế (Hàn Quốc)", 
    "Mã Bưu Chính", 
    "Vĩ độ (Lat)", 
    "Kinh độ (Long)", 
    "Trạng Thái Gán Bản Đồ",
    "Ghi Chú Phân Tích"
]

# Style definitions
font_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
font_data = Font(name="Segoe UI", size=10, color="000000")
font_alert = Font(name="Segoe UI", size=10, bold=True, color="9C0006")

fill_title = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
fill_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
fill_zebra = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
fill_alert_bg = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

thin_side = Side(border_style="thin", color="D9D9D9")
border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# Title block
ws.merge_cells("A1:I1")
title_cell = ws["A1"]
title_cell.value = "DANH SÁCH TRƯỜNG HỌC HÀN QUỐC ĐÃ CHUẨN HÓA GIS & TỌA ĐỘ BẢN ĐỒ"
title_cell.font = font_title
title_cell.fill = fill_title
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 40

# Header row
ws.append([]) # Empty spacer row
ws.row_dimensions[2].height = 10

ws.append(headers)
ws.row_dimensions[3].height = 25
for col_idx in range(1, 10):
    cell = ws.cell(row=3, column=col_idx)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_cell

# Data rows
for idx, s in enumerate(table_data):
    row_idx = idx + 4
    ws.append([
        s['stt'],
        s['name'],
        s['source'],
        s['addr_norm'],
        s['postcode'],
        s['lat'],
        s['lon'],
        s['status'],
        s['notes']
    ])
    ws.row_dimensions[row_idx].height = 22
    
    # Zebra striping
    is_zebra = (idx % 2 == 1)
    
    # Check if there is an alert/correction note
    is_alert = "Đã sửa lỗi" in s['status'] or "Lỗi" in s['status']
    
    for col_idx in range(1, 10):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = font_data
        cell.border = border_cell
        
        # Center align short text columns
        if col_idx in [1, 3, 5, 6, 7, 8]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
        if is_alert:
            cell.fill = fill_alert_bg
            if col_idx == 8 or col_idx == 9:
                cell.font = font_alert
        elif is_zebra:
            cell.fill = fill_zebra

# Set column widths
column_widths = {
    'A': 6,
    'B': 30,
    'C': 18,
    'D': 50,
    'E': 14,
    'F': 14,
    'G': 14,
    'H': 18,
    'I': 35
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

wb.save("DanhSachTruong_Geocoded.xlsx")
print("Saved DanhSachTruong_Geocoded.xlsx successfully!")

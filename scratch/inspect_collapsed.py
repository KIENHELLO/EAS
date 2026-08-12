import openpyxl
import os
import re

# Helper to normalize school names for matching
def clean(text):
    if not text: return ""
    text = str(text).lower()
    text = text.replace('-', '').replace(' ', '').replace('\'', '')
    text = text.replace('đạihọcquốcgia', '').replace('đạihọccônggiáo', '').replace('đạihọcnữsinh', '').replace('đạihọcnữ', '').replace('đạihọc', '')
    text = text.replace('caodẳngkỹthuật', '').replace('caodẳngy', '').replace('caodẳng', '').replace('trường', '')
    patterns = {
        '[àáảãạăằắẳẵặâầấẩẫậ]': 'a',
        '[èéẻẽẹêềếểễệ]': 'e',
        '[ìíỉĩị]': 'i',
        '[òóỏõọôồốổỗộơờớởỡợ]': 'o',
        '[ùúủũụưừứửữự]': 'u',
        '[ỳýỷỹỵ]': 'y',
        'đ': 'd'
    }
    for pattern, repl in patterns.items():
        text = re.sub(pattern, repl, text)
    return text

spelling_map = {
    "seolyong": "seokyeong",
    "seokyeong": "seokyeong",
    "pohang": "postech",
    "postech": "postech",
    "sungkyungwan": "sungkyunkwan",
    "sungkyunkwan": "sungkyunkwan",
    "chungang": "chungang",
    "ewha": "ewha",
    "pusan": "busan",
    "busan": "busan",
    "seoultheological": "seoultheological",
    "kyunghee": "kyunghee",
    "catholic": "catholic",
    "duksung": "duksung",
    "sookmyung": "sookmyung",
    "seoulwomens": "seoulwomens",
    "pukyong": "pukyong",
    "changwon": "changwon",
    "youngsan": "youngsan",
    "yeongsang": "youngsan",
    "doowon": "doowon",
    "jeju": "jeju",
    "osan": "osan",
}

def get_group_key(name):
    c_name = clean(name)
    for k, v in spelling_map.items():
        if k in c_name:
            return v
    return c_name

schools_db = {}
raw_occurrences = [] # (file, name, clean_name, group_key)

def process_name(name, file_source):
    if not name or str(name).strip() == '':
        return
    # Clean check
    val_str = str(name).strip()
    if val_str.isdigit() or val_str.lower() in ['stt', 'tên trường', 'ten truong', 'tên trường học', 'trường', 'danh sách', 'khu vực', 'miền']:
        return
    if 'danh sách' in val_str.lower() or 'stt' in val_str.lower():
        return
        
    g_key = get_group_key(name)
    raw_occurrences.append((file_source, name, clean(name), g_key))
    if g_key not in schools_db:
        schools_db[g_key] = []
    schools_db[g_key].append((file_source, name))

# Read diachitop1%
if os.path.exists("diachitop1%.xlsx"):
    wb = openpyxl.load_workbook("diachitop1%.xlsx", data_only=True)
    for row in list(wb.active.iter_rows(values_only=True))[3:]:
        process_name(row[2], "diachitop1%.xlsx")

# Read diachitop2%
if os.path.exists("diachitop2%.xlsx"):
    wb = openpyxl.load_workbook("diachitop2%.xlsx", data_only=True)
    for row in list(wb.active.iter_rows(values_only=True))[4:]:
        process_name(row[2], "diachitop2%.xlsx")

# Read diachitop3%
if os.path.exists("diachitop3%.xlsx"):
    wb = openpyxl.load_workbook("diachitop3%.xlsx", data_only=True)
    for row in list(wb.active.iter_rows(values_only=True))[2:]:
        process_name(row[1], "diachitop3%.xlsx")

# Now check which group keys have multiple different names representing DIFFERENT schools!
print("Groups with multiple distinct original names:")
for g_key, items in sorted(schools_db.items()):
    unique_names_in_group = list(set([item[1] for item in items]))
    if len(unique_names_in_group) > 1:
        print(f"\nGroup: '{g_key}'")
        for file_source, name in items:
            print(f"  - [{file_source}] {name}")

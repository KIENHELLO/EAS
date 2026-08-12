import openpyxl
import os

files = ['diachitop1%.xlsx', 'diachitop2%.xlsx', 'diachitop3%.xlsx']

for f in files:
    if os.path.exists(f):
        wb = openpyxl.load_workbook(f, read_only=True)
        sheet = wb.active
        print(f"File: {f}")
        print(f"Sheet Name: {sheet.title}")
        print(f"Max Row: {sheet.max_row}")
        
        # Read first 3 rows
        rows = list(sheet.iter_rows(values_only=True))
        print("Header and first few rows:")
        for r in rows[:4]:
            print(r)
        print("-" * 50)
    else:
        print(f"File {f} does not exist.")

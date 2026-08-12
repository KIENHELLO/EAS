import openpyxl
import os

files = ["TOP1%.xlsx", "TO2%.xlsx", "top3.xlsx", "TruongNoTOPIK.xlsx", "danhsachtruonghanche.xlsx", "diachitop1%.xlsx", "diachitop2%.xlsx", "diachitop3%.xlsx"]

for f_name in files:
    if os.path.exists(f_name):
        try:
            wb = openpyxl.load_workbook(f_name, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for r_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    row_str = str(row).lower()
                    if 'kaist' in row_str or '과학기술원' in row_str or 'korea advanced' in row_str:
                        print(f"Found match in {f_name} | Sheet: {sheet_name} | Row: {r_idx}")
                        print("  Row values:", row[:5])
        except Exception as e:
            print(f"Error reading {f_name}: {e}")

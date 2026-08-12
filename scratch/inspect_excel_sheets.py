import openpyxl

wb = openpyxl.load_workbook('MASTER_DATABASE_DAI_HOC_CAO_DANG_HAN_QUOC_2026.xlsx', data_only=True)
for name in wb.sheetnames:
    sheet = wb[name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        print(f"Sheet '{name}' is empty")
        continue
    print(f"Sheet '{name}':")
    print("  Headers:", rows[0])
    print("  Row 1:", rows[1] if len(rows) > 1 else "None")
    print("  Row 2:", rows[2] if len(rows) > 2 else "None")
    print(f"  Total rows: {len(rows)}")

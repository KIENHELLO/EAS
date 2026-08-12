import openpyxl
import os
import re
import urllib.parse
import json

def extract_lat_long(url):
    if not url or not isinstance(url, str):
        return None, None
    match1 = re.search(r'!3d(-?[0-9.]+)!4d(-?[0-9.]+)', url)
    if match1:
        return float(match1.group(1)), float(match1.group(2))
    match2 = re.search(r'/@(-?[0-9.]+),(-?[0-9.]+)', url)
    if match2:
        return float(match2.group(1)), float(match2.group(2))
    try:
        parsed = urllib.parse.urlparse(url)
        qs = urllib.parse.parse_qs(parsed.query)
        if 'q' in qs:
            val = qs['q'][0]
            match3 = re.match(r'(-?[0-9.]+),(-?[0-9.]+)', val)
            if match3:
                return float(match3.group(1)), float(match3.group(2))
    except Exception:
        pass
    return None, None

# Normalize text for matching
def clean(text):
    if not text: return ""
    text = str(text).lower()
    text = text.replace('-', '').replace(' ', '').replace('\'', '')
    text = text.replace('đạihọcquốcgia', '').replace('đạihọccônggiáo', '').replace('đạihọcnữsinh', '').replace('đạihọcnữ', '').replace('đạihọc', '')
    text = text.replace('caodẳngkỹthuật', '').replace('caodẳngy', '').replace('caodẳng', '').replace('trường', '')
    patterns = {
        '[àáảãạăằắẳẵặâầấẩẫậ]': 'a',
        '[èéẻẽẹêềếểễệ]': 'e',
        '[ìíỉĩị]': 'i',
        '[òóỏõọôồốổỗộơờớởỡợ]': 'o',
        '[ùúủũụưừứửữự]': 'u',
        '[ỳýỷỹỵ]': 'y',
        'đ': 'd'
    }
    for pattern, repl in patterns.items():
        text = re.sub(pattern, repl, text)
    return text

# Manual spelling overrides to link names across files
spelling_map = {
    "seolyong": "seokyeong",
    "seokyeong": "seokyeong",
    "pohang": "postech",
    "postech": "postech",
    "sungkyungwan": "sungkyunkwan",
    "sungkyunkwan": "sungkyunkwan",
    "chungang": "chungang",
    "ewha": "ewha",
    "pusan": "busan",
    "busan": "busan",
    "seoultheological": "seoultheological",
    "kyunghee": "kyunghee",
    "catholic": "catholic",
    "duksung": "duksung",
    "sookmyung": "sookmyung",
    "seoulwomens": "seoulwomens",
    "pukyong": "pukyong",
    "changwon": "changwon",
    "youngsan": "youngsan",
    "yeongsang": "youngsan",
    "doowon": "doowon",
    "jeju": "jeju",
    "osan": "osan",
}

def get_group_key(name):
    c_name = clean(name)
    # Check overrides
    for k, v in spelling_map.items():
        if k in c_name:
            return v
    return c_name

# Unique school repository
schools_db = {}

def get_or_create_school(name, file_source):
    g_key = get_group_key(name)
    if not g_key:
        return None
        
    if g_key in schools_db:
        return schools_db[g_key]
        
    is_college_known = any(k in g_key for k in ['bucheon', 'ydongnam', 'jangan', 'kimpo', 'kyungmin', 'shingu', 'doowon', 'tongwon', 'songho', 'gangwonstate', 'masan', 'ajoumotor', 'ajou_motor', 'gangneungyeongdong', 'sangji_catholic', 'cheongam', 'yeungjin'])
    # Create new
    school_entry = {
        "original_names": [name],
        "name_vi": name, # default name
        "file_sources": [file_source],
        "is_top1": False,
        "is_top2": False,
        "is_top3": False,
        "is_no_topik": False,
        "is_restricted": False,
        "is_college": is_college_known,
        "region": None,
        "tuition_desc": None,
        "dorm_desc": None,
        "admission_conditions": None,
        "featured_majors": None,
        "notes": None,
        "maps_link": None,
        "lat": None,
        "lon": None,
        "address_ko": None,
        "postcode": None
    }
    schools_db[g_key] = school_entry
    return school_entry

# Load Address files (which contain maps_link and coordinates)
# TOP 1% Address
if os.path.exists("diachitop1%.xlsx"):
    wb = openpyxl.load_workbook("diachitop1%.xlsx", data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    for r in rows[3:]:
        name = r[2]
        if not name or 'TÊN TRƯỜNG' in str(name) or str(name).strip() == '':
            continue
        s = get_or_create_school(name, "diachitop1%.xlsx")
        if s:
            s["is_top1"] = True
            s["region"] = r[3]
            s["featured_majors"] = r[4]
            s["tuition_desc"] = r[5]
            s["dorm_desc"] = r[6]
            s["admission_conditions"] = r[7]
            s["notes"] = r[10]
            s["maps_link"] = r[12]
            lat, lon = extract_lat_long(s["maps_link"])
            s["lat"] = lat
            s["lon"] = lon

# TOP 2% Address
if os.path.exists("diachitop2%.xlsx"):
    wb = openpyxl.load_workbook("diachitop2%.xlsx", data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    is_college_section = False
    for r in rows[4:]:
        # Check if the row marks a section
        if r[0] and isinstance(r[0], str):
            r0_clean = r[0].strip().upper()
            if 'HỆ CAO ĐẲNG' in r0_clean:
                is_college_section = True
                continue
            if 'HỆ CAO HỌC' in r0_clean:
                is_college_section = False
                continue
            
        name = r[2]
        if not name or 'TÊN TRƯỜNG' in str(name) or str(name).strip() == '' or "MIỀN" in str(name) or "KHU VỰC" in str(name):
            continue
        s = get_or_create_school(name, "diachitop2%.xlsx")
        if s:
            s["is_top2"] = True
            if is_college_section:
                s["is_college"] = True
            s["region"] = r[3]
            s["featured_majors"] = r[4]
            s["tuition_desc"] = r[5]
            s["dorm_desc"] = r[6]
            s["admission_conditions"] = r[7]
            s["maps_link"] = r[11]
            lat, lon = extract_lat_long(s["maps_link"])
            s["lat"] = lat
            s["lon"] = lon

# TOP 3% Address
if os.path.exists("diachitop3%.xlsx"):
    wb = openpyxl.load_workbook("diachitop3%.xlsx", data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    for r in rows[2:]:
        name = r[1]
        if not name or 'TÊN TRƯỜNG' in str(name) or str(name).strip() == '':
            continue
        s = get_or_create_school(name, "diachitop3%.xlsx")
        if s:
            s["is_top3"] = True
            s["region"] = r[3]
            s["featured_majors"] = r[4]
            s["tuition_desc"] = r[5]
            s["dorm_desc"] = r[6]
            s["admission_conditions"] = r[7]
            s["maps_link"] = r[10]
            lat, lon = extract_lat_long(s["maps_link"])
            s["lat"] = lat
            s["lon"] = lon

# Load other sheets to update flags and names
# TOP1%.xlsx (classification)
if os.path.exists("TOP1%.xlsx"):
    wb = openpyxl.load_workbook("TOP1%.xlsx", data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    for r in rows[3:]:
        name = r[2]
        if not name or 'TÊN TRƯỜNG' in str(name) or str(name).strip() == '':
            continue
        s = get_or_create_school(name, "TOP1%.xlsx")
        if s:
            s["is_top1"] = True
            if "TOP1%.xlsx" not in s["file_sources"]: s["file_sources"].append("TOP1%.xlsx")
            if name not in s["original_names"]: s["original_names"].append(name)
            s["region"] = s["region"] or r[3]
            s["featured_majors"] = s["featured_majors"] or r[4]
            s["tuition_desc"] = s["tuition_desc"] or r[5]
            s["dorm_desc"] = s["dorm_desc"] or r[6]
            s["admission_conditions"] = s["admission_conditions"] or r[7]
            s["notes"] = s["notes"] or r[10]
            s["maps_link"] = s["maps_link"] or r[12]
            if s["maps_link"] and (not s.get("lat") or not s.get("lon")):
                lat, lon = extract_lat_long(s["maps_link"])
                s["lat"] = lat or s.get("lat")
                s["lon"] = lon or s.get("lon")

# TO2%.xlsx (classification)
if os.path.exists("TO2%.xlsx"):
    wb = openpyxl.load_workbook("TO2%.xlsx", data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    is_college_section = False
    for r in rows[4:]:
        # Check if the row marks a section
        if r[0] and isinstance(r[0], str):
            r0_clean = r[0].strip().upper()
            if 'HỆ CAO ĐẲNG' in r0_clean:
                is_college_section = True
                continue
            if 'HỆ CAO HỌC' in r0_clean:
                is_college_section = False
                continue
            
        name = r[2]
        if not name or 'TÊN TRƯỜNG' in str(name) or str(name).strip() == '' or "MIỀN" in str(name) or "KHU VỰC" in str(name):
            continue
        s = get_or_create_school(name, "TO2%.xlsx")
        if s:
            s["is_top2"] = True
            if is_college_section:
                s["is_college"] = True
            if "TO2%.xlsx" not in s["file_sources"]: s["file_sources"].append("TO2%.xlsx")
            if name not in s["original_names"]: s["original_names"].append(name)
            s["region"] = s["region"] or r[3]
            s["featured_majors"] = s["featured_majors"] or r[4]
            s["tuition_desc"] = s["tuition_desc"] or r[5]
            s["dorm_desc"] = s["dorm_desc"] or r[6]
            s["admission_conditions"] = s["admission_conditions"] or r[7]

# top3.xlsx (classification)
if os.path.exists("top3.xlsx"):
    wb = openpyxl.load_workbook("top3.xlsx", data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    for r in rows[2:]:
        name = r[1]
        if not name or 'TÊN TRƯỜNG' in str(name) or str(name).strip() == '':
            continue
        s = get_or_create_school(name, "top3.xlsx")
        if s:
            s["is_top3"] = True
            if "top3.xlsx" not in s["file_sources"]: s["file_sources"].append("top3.xlsx")
            if name not in s["original_names"]: s["original_names"].append(name)
            s["region"] = s["region"] or r[3]
            s["featured_majors"] = s["featured_majors"] or r[4]
            s["tuition_desc"] = s["tuition_desc"] or r[5]
            s["dorm_desc"] = s["dorm_desc"] or r[6]
            s["admission_conditions"] = s["admission_conditions"] or r[7]

# TruongNoTOPIK.xlsx
if os.path.exists("TruongNoTOPIK.xlsx"):
    wb = openpyxl.load_workbook("TruongNoTOPIK.xlsx", data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    for r in rows[1:]:
        name = r[1]
        if not name or str(name).strip() == '':
            continue
        # Strip English name in parens from name
        clean_no_topik_name = re.sub(r'\(.*\)', '', str(name)).strip()
        s = get_or_create_school(clean_no_topik_name, "TruongNoTOPIK.xlsx")
        if s:
            s["is_no_topik"] = True
            if "TruongNoTOPIK.xlsx" not in s["file_sources"]: s["file_sources"].append("TruongNoTOPIK.xlsx")
            if name not in s["original_names"]: s["original_names"].append(name)

# danhsachtruonghanche.xlsx
if os.path.exists("danhsachtruonghanche.xlsx"):
    wb = openpyxl.load_workbook("danhsachtruonghanche.xlsx", data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    for r in rows[1:]:
        name = r[1]
        if not name or str(name).strip() == '':
            continue
        s = get_or_create_school(name, "danhsachtruonghanche.xlsx")
        if s:
            s["is_restricted"] = True
            if "danhsachtruonghanche.xlsx" not in s["file_sources"]: s["file_sources"].append("danhsachtruonghanche.xlsx")
            if name not in s["original_names"]: s["original_names"].append(name)

# Let's count totals
print(f"Total compiled unique schools in our unified database: {len(schools_db)}")
print(f"  Top 1% count: {sum(1 for s in schools_db.values() if s['is_top1'])}")
print(f"  Top 2% count: {sum(1 for s in schools_db.values() if s['is_top2'])}")
print(f"  Top 3% count: {sum(1 for s in schools_db.values() if s['is_top3'])}")
print(f"  No TOPIK count: {sum(1 for s in schools_db.values() if s['is_no_topik'])}")
print(f"  Restricted count: {sum(1 for s in schools_db.values() if s['is_restricted'])}")

# Save to scratch file
with open("scratch/unified_schools.json", "w", encoding="utf-8") as f:
    json.dump(list(schools_db.values()), f, ensure_ascii=False, indent=2)
print("Saved scratch/unified_schools.json successfully!")

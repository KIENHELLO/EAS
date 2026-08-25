import openpyxl
import os
import re
import json

# 1. Load database
def load_db(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    prefix = "export const universities ="
    if prefix in content:
        json_str = content.replace(prefix, '').replace(';', '').strip()
        return json.loads(json_str)
    return []

db = load_db('src/data/universities.js')

# 2. Helpers for normalization
def clean_name(name):
    if not name:
        return ""
    name = str(name).lower()
    name = re.sub(r'\(.*?\)', '', name)
    name = re.sub(r'\[.*?\]', '', name)
    name = re.sub(r'[đĐ]', 'd', name)
    name = re.sub(r'[àáạảãâầấậẩẫăằắặẳẵ]', 'a', name)
    name = re.sub(r'[èéẹẻẽêềếệểễ]', 'e', name)
    name = re.sub(r'[ìíịỉĩ]', 'i', name)
    name = re.sub(r'[òóọỏõôồốộổỗơờớợởỡ]', 'o', name)
    name = re.sub(r'[ùúụủũưừứựửữ]', 'u', name)
    name = re.sub(r'[ỳýỵỷỹ]', 'y', name)
    name = re.sub(r'[^a-z0-9\uac00-\ud7a3\s]', '', name)
    return " ".join(name.split())

def extract_hangeul(text):
    if not text:
        return ""
    matches = re.findall(r'[\uac00-\ud7a3]+', text)
    return "".join(matches)

# 3. Robust function to find header row index
def find_header_row(rows):
    for r_idx, row in enumerate(rows):
        row_str = [str(cell).lower() for cell in row if cell is not None]
        if any('trường' in s or 'name' in s or 'school' in s or 'tên' in s for s in row_str):
            return r_idx
    return 0

# 4. Read detail Excel files
excel_files = ['TOP1%.xlsx', 'TO2%.xlsx', 'top3.xlsx', 'diachitop1%.xlsx', 'diachitop2%.xlsx', 'diachitop3%.xlsx']
excel_data = {}

for f_name in excel_files:
    if not os.path.exists(f_name):
        continue
    wb = openpyxl.load_workbook(f_name, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        continue
    
    header_row_idx = find_header_row(rows)
    headers = [str(h).strip().lower() if h else "" for h in rows[header_row_idx]]
    
    name_idx = -1
    cond_idx = -1
    
    for idx, h in enumerate(headers):
        if any(x in h for x in ['trường', 'tên', 'name', 'school']):
            name_idx = idx
        elif any(x in h for x in ['điều kiện', 'tuyển sinh', 'gpa', 'cond']):
            cond_idx = idx
            
    if name_idx == -1: name_idx = 1
    if cond_idx == -1: cond_idx = 7
    
    # Process rows after header
    for row in rows[header_row_idx + 1:]:
        if not any(x is not None for x in row):
            continue
        school_name = row[name_idx] if name_idx < len(row) else None
        if not school_name or any(x in str(school_name).lower() for x in ['tên trường', 'school name', 'stt']):
            continue
            
        cond = str(row[cond_idx]).strip() if cond_idx != -1 and cond_idx < len(row) and row[cond_idx] is not None else ""
        if cond.lower() == 'n/a' or cond == 'None' or cond == 'none':
            cond = ""
            
        c_name = clean_name(school_name)
        h_name = extract_hangeul(str(school_name))
        
        if c_name:
            excel_data[c_name] = (cond, f_name, school_name)
        if h_name:
            excel_data[h_name] = (cond, f_name, school_name)

print(f"Loaded {len(excel_data)} reference points from Excel files.")

# 5. Audit DB vs Excel
mismatches = 0
matches = 0
not_found_in_excel = 0

for u in db:
    u_ko_hang = extract_hangeul(u.get('name_ko', ''))
    u_vi_clean = clean_name(u.get('name_vi', ''))
    u_en_clean = clean_name(u.get('name_en', ''))
    u_ko_clean = clean_name(u.get('name_ko', ''))
    
    ref = None
    if u_ko_hang and u_ko_hang in excel_data:
        ref = excel_data[u_ko_hang]
    elif u_ko_clean and u_ko_clean in excel_data:
        ref = excel_data[u_ko_clean]
    elif u_vi_clean and u_vi_clean in excel_data:
        ref = excel_data[u_vi_clean]
    elif u_en_clean and u_en_clean in excel_data:
        ref = excel_data[u_en_clean]
        
    db_cond = u.get('admission_conditions', '')
    if db_cond is None: db_cond = ""
    db_cond = db_cond.strip()
    
    if ref:
        excel_cond, file_src, excel_name = ref
        excel_cond = excel_cond.strip()
        
        if db_cond != excel_cond:
            mismatches += 1
            print(f"MISMATCH: '{u['name_vi'].replace(chr(10), ' ')}' ({u['id']})")
            print(f"  Source file: {file_src} (School: '{excel_name}')")
            print(f"  Database has: '{db_cond}'")
            print(f"  Excel has:    '{excel_cond}'")
        else:
            matches += 1
    else:
        not_found_in_excel += 1

print("\nAudit Summary:")
print(f"Total schools in DB: {len(db)}")
print(f"Matched & Verified (DB equals Excel): {matches}")
print(f"Mismatches found: {mismatches}")
print(f"Schools not in detail Excel lists (using default placeholder): {not_found_in_excel}")

import openpyxl
import os

files = ['TOP1%.xlsx', 'TO2%.xlsx', 'top3.xlsx', 'diachitop1%.xlsx', 'diachitop2%.xlsx', 'diachitop3%.xlsx']
for f in files:
    if os.path.exists(f):
        wb = openpyxl.load_workbook(f, data_only=True)
        print(f"File '{f}': Sheets: {wb.sheetnames}")
        sheet = wb[wb.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
        if rows:
            print("  Headers:", rows[0])
            non_empty_rows = [r for r in rows if any(x is not None for x in r)]
            print(f"  Total non-empty rows: {len(non_empty_rows)}")
            if len(non_empty_rows) > 1:
                print("  Sample row 1:", non_empty_rows[1])
            if len(non_empty_rows) > 2:
                print("  Sample row 2:", non_empty_rows[2])
    else:
        print(f"File '{f}' does not exist")

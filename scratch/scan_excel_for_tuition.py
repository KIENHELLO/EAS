import openpyxl

wb = openpyxl.load_workbook('MASTER_DATABASE_DAI_HOC_CAO_DANG_HAN_QUOC_2026.xlsx', data_only=True)
keywords = ['học phí', 'tuition', '등록금', '학비', 'fee', 'cost', 'money', 'krw', 'won']

for name in wb.sheetnames:
    sheet = wb[name]
    found = False
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        for c_idx, val in enumerate(row):
            if val is not None:
                val_str = str(val).lower()
                if any(k in val_str for k in keywords):
                    print(f"Sheet '{name}', Cell ({r_idx+1},{c_idx+1}): {val}")
                    found = True
    if not found:
        print(f"Sheet '{name}': No keywords found.")

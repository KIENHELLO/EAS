import openpyxl
import os
import re
import json

# 1. Load database
def load_db(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    prefix = "export const universities ="
    if prefix in content:
        json_str = content.replace(prefix, '').replace(';', '').strip()
        return json.loads(json_str)
    return []

db = load_db('src/data/universities.js')
print("Loaded schools from database:", len(db))

# 2. Helper to clean and normalize names for fuzzy matching
def clean_name(name):
    if not name:
        return ""
    name = str(name).lower()
    name = re.sub(r'\(.*?\)', '', name)
    name = re.sub(r'\[.*?\]', '', name)
    # Remove accents/diacritics from Vietnamese text for robust matching
    name = re.sub(r'[đĐ]', 'd', name)
    name = re.sub(r'[àáạảãâầấậẩẫăằắặẳẵ]', 'a', name)
    name = re.sub(r'[èéẹẻẽêềếệểễ]', 'e', name)
    name = re.sub(r'[ìíịỉĩ]', 'i', name)
    name = re.sub(r'[òóọỏõôồốộổỗơờớợởỡ]', 'o', name)
    name = re.sub(r'[ùúụủũưừứựửữ]', 'u', name)
    name = re.sub(r'[ỳýỵỷỹ]', 'y', name)
    # Strip non-alphanumeric except spaces
    name = re.sub(r'[^a-z0-9\uac00-\ud7a3\s]', '', name)
    return " ".join(name.split())

def extract_hangeul(text):
    if not text:
        return ""
    matches = re.findall(r'[\uac00-\ud7a3]+', text)
    return "".join(matches)

# 3. Read detail Excel files
excel_files = ['TOP1%.xlsx', 'TO2%.xlsx', 'top3.xlsx', 'diachitop1%.xlsx', 'diachitop2%.xlsx', 'diachitop3%.xlsx']
excel_data = []

for f_name in excel_files:
    if not os.path.exists(f_name):
        print(f"File {f_name} does not exist")
        continue
    wb = openpyxl.load_workbook(f_name, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        continue
    
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    if len(rows) > 1 and all(x is None for x in rows[0]):
        # Backup headers if row 0 is empty
        headers = [str(h).strip().lower() if h else "" for h in rows[1]]
        start_row = 2
    else:
        start_row = 1
        
    name_idx = -1
    invoice_idx = -1
    ktx_idx = -1
    cond_idx = -1
    majors_idx = -1
    
    for idx, h in enumerate(headers):
        if any(x in h for x in ['trường', 'tên', 'name', 'school']):
            name_idx = idx
        elif any(x in h for x in ['invoice', 'học phí', 'tuition']):
            invoice_idx = idx
        elif any(x in h for x in ['ktx', 'ký túc', 'dorm']):
            ktx_idx = idx
        elif any(x in h for x in ['điều kiện', 'tuyển sinh', 'gpa', 'cond']):
            cond_idx = idx
        elif any(x in h for x in ['ngành', 'thế mạnh', 'major']):
            majors_idx = idx
            
    # Default indexes if search failed
    if name_idx == -1: name_idx = 1
    if invoice_idx == -1: invoice_idx = 5
    if ktx_idx == -1: ktx_idx = 6
    if cond_idx == -1: cond_idx = 7
    
    print(f"File '{f_name}': name_idx={name_idx}, invoice_idx={invoice_idx}, ktx_idx={ktx_idx}, cond_idx={cond_idx}")
    
    for row in rows[start_row:]:
        if not any(x is not None for x in row):
            continue
        school_name = row[name_idx] if name_idx < len(row) else None
        if not school_name or any(x in str(school_name).lower() for x in ['tên trường', 'school name', 'stt']):
            continue
            
        invoice = str(row[invoice_idx]).strip() if invoice_idx != -1 and invoice_idx < len(row) and row[invoice_idx] is not None else None
        ktx = str(row[ktx_idx]).strip() if ktx_idx != -1 and ktx_idx < len(row) and row[ktx_idx] is not None else None
        cond = str(row[cond_idx]).strip() if cond_idx != -1 and cond_idx < len(row) and row[cond_idx] is not None else None
        majors = str(row[majors_idx]).strip() if majors_idx != -1 and majors_idx < len(row) and row[majors_idx] is not None else None
        
        c_name = clean_name(school_name)
        if not c_name:
            continue
            
        excel_data.append({
            'file': f_name,
            'raw_name': school_name,
            'clean_name': c_name,
            'hangeul': extract_hangeul(str(school_name)),
            'invoice': invoice,
            'ktx': ktx,
            'cond': cond,
            'majors': majors
        })

print(f"Loaded {len(excel_data)} detail records from Excel files.")

# 4. Perform matching and merge details
matched_count = 0
for u in db:
    u_ko_hang = extract_hangeul(u.get('name_ko', ''))
    u_vi_clean = clean_name(u.get('name_vi', ''))
    u_en_clean = clean_name(u.get('name_en', ''))
    u_ko_clean = clean_name(u.get('name_ko', ''))
    
    # Ignore placeholders or empty
    if not u_vi_clean or u_vi_clean == 'n a':
        continue
        
    best_match = None
    
    for item in excel_data:
        it_name = item['clean_name']
        it_hang = item['hangeul']
        
        # Match 1: Exact Hangeul match
        if u_ko_hang and it_hang and u_ko_hang == it_hang:
            best_match = item
            break
            
        # Match 2: Cleaned names are identical
        if u_ko_clean and u_ko_clean == it_name:
            best_match = item
            break
        if u_vi_clean and u_vi_clean == it_name:
            best_match = item
            break
        if u_en_clean and u_en_clean == it_name:
            best_match = item
            break
            
        # Match 3: Cleaned names contains or is contained in (with min length 5)
        if len(u_vi_clean) >= 5 and len(it_name) >= 5:
            if u_vi_clean in it_name or it_name in u_vi_clean:
                best_match = item
                break
        if len(u_en_clean) >= 5 and len(it_name) >= 5:
            if u_en_clean in it_name or it_name in u_en_clean:
                best_match = item
                break

    if best_match:
        matched_count += 1
        print(f"Matched school: '{u['name_vi'].replace(chr(10), ' ')}' -> Excel '{best_match['raw_name']}' ({best_match['file']})")
        
        # Merge fields
        if best_match['cond'] and str(best_match['cond']).strip().lower() != 'n/a' and str(best_match['cond']).strip() != "":
            u['admission_conditions'] = best_match['cond']
            print(f"  -> admission_conditions: {best_match['cond']}")
            
        if best_match['ktx'] and str(best_match['ktx']).strip().lower() != 'n/a' and str(best_match['ktx']).strip() != "":
            u['dorm_fee_desc'] = best_match['ktx']
            
        if best_match['majors'] and str(best_match['majors']).strip().lower() != 'n/a' and str(best_match['majors']).strip() != "":
            u['featured_majors'] = best_match['majors']
            
        if best_match['invoice'] and str(best_match['invoice']).strip().lower() != 'n/a' and str(best_match['invoice']).strip() != "":
            u['invoice_details'] = best_match['invoice']

print(f"Successfully matched and updated {matched_count} out of {len(db)} schools.")

# 5. Save updated database to both projects
def save_db(db_data, file_path):
    new_content = f"export const universities = {json.dumps(db_data, indent=2, ensure_ascii=False)};\n"
    with open(file_path, 'w', encoding='utf-8') as f:
        f.write(new_content)
    print(f"Saved database to {file_path}")

save_db(db, 'src/data/universities.js')
save_db(db, 'kr-unituition-next/src/data/universities.js')

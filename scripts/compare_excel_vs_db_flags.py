import json
import openpyxl
import re

with open('scratch/unis.json', 'r', encoding='utf-8') as f:
    unis = json.load(f)

def normalize_name(s):
    if not s: return ''
    s = str(s).lower().strip()
    s = re.sub(r'[àáạảãâầấậẩẫăằắặẳẵ]', 'a', s)
    s = re.sub(r'[èéẹẻẽêềếệểễ]', 'e', s)
    s = re.sub(r'[ìíịỉĩ]', 'i', s)
    s = re.sub(r'[òóọỏõôồốộổỗơờớợởỡ]', 'o', s)
    s = re.sub(r'[ùúụủũưừứựửữ]', 'u', s)
    s = re.sub(r'[ỳýỵỷỹ]', 'y', s)
    s = re.sub(r'[đ]', 'd', s)
    s = re.sub(r'[\(\)\-\,\.\_\/]', ' ', s)
    words = [w for w in s.split() if w not in ['dai', 'hoc', 'cao', 'dang', 'vien', 'trung', 'cap', 'school', 'university', 'college', 'top', '1%', '2%', '3%']]
    return " ".join(words)

aliases = {
    "sungkyungwan": "skku", "sungkyunkwan": "skku", "chungang": "cau", "chung ang": "cau",
    "pohang": "postech", "busan": "pusan", "seoultech": "seoultech", "sungshin": "sungshin",
    "sookmyung": "sookmyung", "duksung": "duksung", "ewha": "ewha", "khu": "khu", "kyunghee": "khu",
    "gangwon": "kangwon", "kangwon": "kangwon", "soonchunhyang": "uni_8", "hallym": "uni_9",
    "myongji": "uni_10", "sangmyung": "uni_11", "gachon": "uni_1", "dankook": "uni_2",
    "chosun": "uni_3", "donga": "uni_4", "pukyong": "uni_5", "ulsan": "uni_6", "changwon": "uni_7",
    "kyungil": "kyungil", "dong eui": "dongeui", "dongeui": "dongeui", "yong in": "yongin",
    "yongin": "yongin", "dongduk": "dongduk", "mokpo": "mock_uni_120", "sangji": "sangji",
    "chodang": "chodang", "halla": "halla", "keimyung": "keimyung"
}

def find_best_db_match(excel_name):
    norm_ex = normalize_name(excel_name)
    for alias_k, target_id in aliases.items():
        if alias_k in norm_ex or norm_ex == alias_k:
            for u in unis:
                if u['id'] == target_id: return u
    for u in unis:
        norm_vi = normalize_name(u['name_vi'])
        norm_en = normalize_name(u.get('name_en', ''))
        norm_ko = normalize_name(u.get('name_ko', ''))
        if norm_vi and (norm_vi in norm_ex or norm_ex in norm_vi): return u
        if norm_en and (norm_en in norm_ex or norm_ex in norm_en): return u
        if norm_ko and (norm_ko in norm_ex or norm_ex in norm_ko): return u
    return None

def extract_excel_names(filename):
    wb = openpyxl.load_workbook(filename, data_only=True)
    raw_list = []
    for sname in wb.sheetnames:
        sheet = wb[sname]
        for row in sheet.iter_rows(values_only=True):
            if not row: continue
            row_cells = [str(c).strip() for c in row if c is not None and str(c).strip() != '']
            for cell_text in row_cells:
                if any(kw in cell_text for kw in ["Đại học", "Cao đẳng", "Viện", "ĐH", "CĐ", "University", "College"]):
                    if not any(h in cell_text.lower() for h in ["danh sách", "tên trường", "stt", "bảng", "loại trường"]):
                        if cell_text not in raw_list:
                            raw_list.append(cell_text)
                            break
    return raw_list

top1_excel = extract_excel_names('TOP1%.xlsx')
top2_excel = extract_excel_names('TO2%.xlsx')
top3_excel = extract_excel_names('top3.xlsx')
no_topik_excel = extract_excel_names('TruongNoTOPIK.xlsx')
restricted_excel = extract_excel_names('danhsachtruonghanche.xlsx')

print(f"============================================================")
print(f"   ĐỐI CHIẾU CỜ (FLAGS) TRONG EXCEL VS UNIVERSITIES.JS    ")
print(f"============================================================\n")

# 1. Check TOP 1%
top1_matched = [find_best_db_match(name) for name in top1_excel if find_best_db_match(name)]
top1_ids = set([u['id'] for u in top1_matched])
print(f"📌 TOP 1% EXCEL: Tìm thấy {len(top1_excel)} trường trong Excel -> Khớp {len(top1_ids)} ID trường trong DB.")
db_top1_count = len([u for u in unis if u.get('top_1_percent') == True or u.get('accept_gdtx') == 'top1'])
print(f"   Website hiển thị TOP 1%: {db_top1_count} trường.")

# 2. Check TOP 2%
top2_matched = [find_best_db_match(name) for name in top2_excel if find_best_db_match(name)]
top2_ids = set([u['id'] for u in top2_matched])
print(f"\n📌 TOP 2% EXCEL: Tìm thấy {len(top2_excel)} trường trong Excel -> Khớp {len(top2_ids)} ID trường trong DB.")
db_top2_count = len([u for u in unis if u.get('accept_gdtx') == 'top2'])
print(f"   Website hiển thị TOP 2%: {db_top2_count} trường.")

# 3. Check TOP 3%
top3_matched = [find_best_db_match(name) for name in top3_excel if find_best_db_match(name)]
top3_ids = set([u['id'] for u in top3_matched])
print(f"\n📌 TOP 3% EXCEL: Tìm thấy {len(top3_excel)} trường trong Excel -> Khớp {len(top3_ids)} ID trường trong DB.")
db_top3_count = len([u for u in unis if u.get('accept_gdtx') == 'top3'])
print(f"   Website hiển thị TOP 3%: {db_top3_count} trường.")

# 4. Check Nợ TOPIK
no_topik_matched = [find_best_db_match(name) for name in no_topik_excel if find_best_db_match(name)]
no_topik_ids = set([u['id'] for u in no_topik_matched])
print(f"\n📌 NỢ TOPIK EXCEL: Tìm thấy {len(no_topik_excel)} trường trong Excel -> Khớp {len(no_topik_ids)} ID trường trong DB.")
db_no_topik_count = len([u for u in unis if u.get('master_no_topik') == True])
print(f"   Website hiển thị Thạc sĩ Nợ TOPIK: {db_no_topik_count} trường.")

# 5. Check Hạn chế Visa
restricted_matched = [find_best_db_match(name) for name in restricted_excel if find_best_db_match(name)]
restricted_ids = set([u['id'] for u in restricted_matched])
print(f"\n📌 HẠN CHẾ VISA EXCEL: Tìm thấy {len(restricted_excel)} trường trong Excel -> Khớp {len(restricted_ids)} ID trường trong DB.")
db_restricted_count = len([u for u in unis if u.get('is_restricted_school') == True])
print(f"   Website hiển thị Hạn chế Visa: {db_restricted_count} trường.")

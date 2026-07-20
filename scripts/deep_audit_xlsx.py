import openpyxl
import json
import re
import subprocess

node_script = """
import { universities } from '../src/data/universities.js';
import fs from 'fs';
fs.writeFileSync('scratch/unis.json', JSON.stringify(universities, null, 2), 'utf-8');
"""
with open('scratch/export_unis.js', 'w', encoding='utf-8') as f:
    f.write(node_script)

subprocess.run(['node', 'scratch/export_unis.js'], check=True)

with open('scratch/unis.json', 'r', encoding='utf-8') as f:
    unis = json.load(f)

print(f"============================================================")
print(f"  BÁO CÁO ĐỐI CHIẾU TOÀN DIỆN VÀ MINH BẠCH VỀ CÁC FILE XLSX ")
print(f"============================================================\n")

print(f"📌 Tổng số trường hiện có trong hệ thống website (universities.js): {len(unis)} trường\n")

def remove_tones(s):
    if not s: return ''
    s = re.sub(r'[àáạảãâầấậẩẫăằắặẳẵ]', 'a', str(s))
    s = re.sub(r'[èéẹẻẽêềếệểễ]', 'e', str(s))
    s = re.sub(r'[ìíịỉĩ]', 'i', str(s))
    s = re.sub(r'[òóọỏõôồốộổỗơờớợởỡ]', 'o', str(s))
    s = re.sub(r'[ùúụủũưừứựửữ]', 'u', str(s))
    s = re.sub(r'[ỳýỵỷỹ]', 'y', str(s))
    s = re.sub(r'[đ]', 'd', str(s))
    return s.lower().strip()

# Analyze each XLSX file row by row
files_info = [
    ("TOP1%.xlsx", "Danh sách trường Đại học TOP 1%"),
    ("TO2%.xlsx", "Danh sách trường Đại học TOP 2%"),
    ("top3.xlsx", "Danh sách trường Đại học TOP 3%"),
    ("TruongNoTOPIK.xlsx", "Danh sách trường Thạc sĩ / Nợ TOPIK"),
    ("danhsachtruonghanche.xlsx", "Danh sách trường Hạn chế cấp Visa")
]

for filename, title in files_info:
    print(f"------------------------------------------------------------")
    print(f"📄 TỆP EXCEL: {filename} ({title})")
    print(f"------------------------------------------------------------")
    
    wb = openpyxl.load_workbook(filename, data_only=True)
    extracted_schools = []
    
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if not row: continue
            row_str = " | ".join([str(c).strip() for c in row if c is not None and str(c).strip() != ''])
            if not row_str: continue
            
            # Find candidate university name cell
            for c in row:
                if c and isinstance(c, str):
                    clean_c = c.strip()
                    # Filter out headers like "Tên trường", "STT", "Học phí"
                    if any(kw in clean_c for kw in ["Đại học", "Cao đẳng", "Viện", "ĐH", "CĐ", "University", "College"]) and len(clean_c) < 60:
                        if not any(hdr in clean_c.lower() for hdr in ["tên trường", "danh sách", "stt", "bảng", "trường đại học"]):
                            if clean_c not in extracted_schools:
                                extracted_schools.append(clean_c)

    print(f"👉 Số lượng trường đọc được từ tệp {filename}: {len(extracted_schools)} trường")
    print(f"📋 Danh sách tên trường trích xuất từ Excel:")
    for idx, name in enumerate(extracted_schools, start=1):
        # Match with universities.js
        norm_name = remove_tones(name)
        matched_uni = None
        for u in unis:
            if remove_tones(u['name_vi']) in norm_name or norm_name in remove_tones(u['name_vi']) or (u.get('name_en') and remove_tones(u['name_en']) in norm_name):
                matched_uni = u
                break
        
        status_str = f"✅ Khớp trong DB: {matched_uni['name_vi']} (ID: {matched_uni['id']}, GDTX: {matched_uni.get('accept_gdtx')}, Top1%: {matched_uni.get('top_1_percent')})" if matched_uni else "❌ Chưa khớp trong DB"
        print(f"   {idx}. {name} --> {status_str}")
    print("\n")

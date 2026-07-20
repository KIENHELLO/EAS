import openpyxl
import json
import re

# 1. Standard Faculty Definitions & Key Major Catalog
FACULTIES = {
    "humanities_social": {
        "name_vi": "Khoa Nhân văn & Khoa học Xã hội",
        "name_ko": "인문사회대학",
        "keywords": ["kinh doanh", "quản trị", "kinh tế", "ngôn ngữ", "truyền thông", "báo chí", "thương mại", "du lịch", "khách sạn", "tâm lý", "sư phạm", "ngoại ngữ", "thần học", "chính trị", "hành chính", "xã hội", "bất động sản", "văn hóa", "quốc tế"]
    },
    "engineering": {
        "name_vi": "Khoa Kỹ thuật & Công nghệ Thông tin",
        "name_ko": "공과대학",
        "keywords": ["kỹ thuật", "công nghệ", "máy tính", "cntt", "it", "ai", "trí tuệ nhân tạo", "điện", "điện tử", "bán dẫn", "cơ khí", "ô tô", "kiến trúc", "xây dựng", "đóng tàu", "phòng cháy", "môi trường"]
    },
    "natural_sciences": {
        "name_vi": "Khoa Khoa học Tự nhiên & Sinh học",
        "name_ko": "자연과학대학",
        "keywords": ["tự nhiên", "sinh học", "hóa học", "toán", "vật lý", "thực phẩm", "nông nghiệp", "dinh dưỡng", "thú y", "y sinh"]
    },
    "arts_sports": {
        "name_vi": "Khoa Nghệ thuật & Thể dục Thể thao",
        "name_ko": "예체능대학",
        "keywords": ["thiết kế", "mỹ thuật", "nghệ thuật", "âm nhạc", "điện ảnh", "sân khấu", "phim", "thẩm mỹ", "làm đẹp", "thể thao", "thời trang", "animation", "hoạt hình", "make-up", "trang điểm", "tóc"]
    },
    "medicine_pharmacy": {
        "name_vi": "Khoa Y - Dược & Khoa học Y tế",
        "name_ko": "의치약학/보건대학",
        "keywords": ["y khoa", "y học", "dược", "điều dưỡng", "vật lý trị liệu", "y tế", "nursing", "phục hồi chức năng", "nha khoa"]
    }
}

# Standard major names dictionary with Korean translations
MAJOR_DICTIONARY = [
    # Humanities & Social
    {"name_vi": "Quản trị Kinh doanh", "name_ko": "경영학과", "cat": "humanities_social", "is_hot": True},
    {"name_vi": "Ngôn ngữ & Văn hóa Hàn Quốc", "name_ko": "한국언어문화학과", "cat": "humanities_social", "is_hot": True},
    {"name_vi": "Truyền thông & Báo chí", "name_ko": "미디어언론학과", "cat": "humanities_social", "is_hot": True},
    {"name_vi": "Thương mại Quốc tế", "name_ko": "국제무역학과", "cat": "humanities_social", "is_hot": True},
    {"name_vi": "Quản trị Du lịch & Khách sạn", "name_ko": "관광호텔경영학과", "cat": "humanities_social", "is_hot": True},
    {"name_vi": "Ngôn ngữ Anh / Biên phiên dịch", "name_ko": "영어영문학과", "cat": "humanities_social", "is_hot": False},
    {"name_vi": "Kinh tế học", "name_ko": "경제학과", "cat": "humanities_social", "is_hot": False},
    {"name_vi": "Tâm lý học", "name_ko": "심리학과", "cat": "humanities_social", "is_hot": False},
    {"name_vi": "Hành chính Công & Chính trị", "name_ko": "행정정치학과", "cat": "humanities_social", "is_hot": False},
    
    # Engineering & IT
    {"name_vi": "Khoa học Máy tính & AI", "name_ko": "컴퓨터공학과", "cat": "engineering", "is_hot": True},
    {"name_vi": "Công nghệ Thông tin (IT)", "name_ko": "정보통신공학과", "cat": "engineering", "is_hot": True},
    {"name_vi": "Kỹ thuật Điện - Điện tử", "name_ko": "전기전자공학과", "cat": "engineering", "is_hot": True},
    {"name_vi": "Kỹ thuật Cơ khí", "name_ko": "기계공학과", "cat": "engineering", "is_hot": True},
    {"name_vi": "Kỹ thuật Ô tô & Xe thông minh", "name_ko": "미래자동차공학과", "cat": "engineering", "is_hot": True},
    {"name_vi": "Kỹ thuật Bán dẫn", "name_ko": "반도체공학과", "cat": "engineering", "is_hot": False},
    {"name_vi": "Kỹ thuật Kiến trúc & Xây dựng", "name_ko": "건축공학과", "cat": "engineering", "is_hot": False},
    
    # Arts & Sports
    {"name_vi": "Thiết kế Đồ họa & Kỹ thuật số", "name_ko": "시각디자인학과", "cat": "arts_sports", "is_hot": True},
    {"name_vi": "Nghệ thuật Thẩm mỹ & Trang điểm (Make-up)", "name_ko": "뷰티미용학과", "cat": "arts_sports", "is_hot": True},
    {"name_vi": "Diễn xuất, Sân khấu & Điện ảnh", "name_ko": "연극영화학과", "cat": "arts_sports", "is_hot": True},
    {"name_vi": "Âm nhạc K-Pop & Vẫn nghệ", "name_ko": "실용음악학과", "cat": "arts_sports", "is_hot": True},
    {"name_vi": "Thiết kế Thời trang", "name_ko": "패션디자인학과", "cat": "arts_sports", "is_hot": False},

    # Natural Sciences
    {"name_vi": "Hóa học & Vật liệu mới", "name_ko": "화학과", "cat": "natural_sciences", "is_hot": False},
    {"name_vi": "Công nghệ Thực phẩm & Dinh dưỡng", "name_ko": "식품영양학과", "cat": "natural_sciences", "is_hot": False},
    {"name_vi": "Sinh học Ứng dụng & Công nghệ Sinh học", "name_ko": "생명공학과", "cat": "natural_sciences", "is_hot": True},

    # Medicine & Pharmacy
    {"name_vi": "Điều dưỡng (Nursing)", "name_ko": "간호학과", "cat": "medicine_pharmacy", "is_hot": True},
    {"name_vi": "Quản lý Y tế & Bệnh viện", "name_ko": "보건행정학과", "cat": "medicine_pharmacy", "is_hot": False},
    {"name_vi": "Vật lý trị liệu & Phục hồi chức năng", "name_ko": "물리치료학과", "cat": "medicine_pharmacy", "is_hot": False},
    {"name_vi": "Dược học", "name_ko": "약학과", "cat": "medicine_pharmacy", "is_hot": False}
]

# Read XLSX files and collect school featured majors
xlsx_files = ["TOP1%.xlsx", "TO2%.xlsx", "top3.xlsx"]
raw_xlsx_map = {}

for filename in xlsx_files:
    filepath = f"d:/EASS/{filename}"
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb.active
    
    name_col = None
    major_col = None
    for r in range(1, 5):
        for c in range(1, sheet.max_column + 1):
            val = str(sheet.cell(row=r, column=c).value or '').strip().upper()
            if 'TRƯỜNG' in val and not name_col:
                name_col = c
            if ('THẾ MẠNH' in val or 'NGÀNH NỔI TRỘI' in val or 'NGÀNH NỔI BẬT' in val) and not major_col:
                major_col = c
                
    if name_col and major_col:
        for r in range(2, sheet.max_row + 1):
            s_name = str(sheet.cell(row=r, column=name_col).value or '').strip()
            s_majors = str(sheet.cell(row=r, column=major_col).value or '').strip()
            if s_name and len(s_name) > 2 and 'STT' not in s_name:
                # Clean school name
                clean_name = re.sub(r'\(.*?\)', '', s_name).strip()
                raw_xlsx_map[clean_name.lower()] = s_majors
                raw_xlsx_map[s_name.lower()] = s_majors

print(f"Collected raw XLSX majors for {len(raw_xlsx_map)} keys.")

# Update universities.js
paths = [
    "d:/EASS/src/data/universities.js",
    "d:/EASS/kr-unituition-next/src/data/universities.js"
]

for path in paths:
    with open(path, "r", encoding="utf-8") as f:
        js_content = f.read()

    match = re.search(r'export const universities = (\[.*\]);', js_content, re.DOTALL)
    if not match:
        print(f"Error reading {path}")
        continue
        
    universities = json.loads(match.group(1))
    print(f"Processing {len(universities)} universities in {path}...")
    
    for u in universities:
        vi_name = u["name_vi"].lower()
        en_name = u["name_en"].lower()
        
        # Match with XLSX extracted text
        found_xlsx_majors = ""
        for k, v in raw_xlsx_map.items():
            if k in vi_name or k in en_name or vi_name in k:
                found_xlsx_majors = v
                break
                
        if found_xlsx_majors and not u.get("featured_majors"):
            u["featured_majors"] = found_xlsx_majors
            
        feat_text = (u.get("featured_majors") or "") + " " + (found_xlsx_majors or "") + " " + u.get("description", "")
        feat_text_lower = feat_text.lower()
        
        # Build majors_detail array
        majors_detail = []
        tuition_dict = u.get("tuition", {})
        
        for cat_key, cat_info in FACULTIES.items():
            # Get semester tuition for this category
            sem_tuition = tuition_dict.get(cat_key) or tuition_dict.get("humanities_social") or 3500000
            
            # Select majors matching dictionary
            faculty_majors = []
            for item in MAJOR_DICTIONARY:
                if item["cat"] == cat_key:
                    # Check if this major is relevant to the school or is a standard major
                    kw_match = any(kw in feat_text_lower for kw in cat_info["keywords"])
                    is_hot = item["is_hot"]
                    
                    # Boost hot flag if explicitly mentioned in featured_majors
                    major_kw_match = any(w in feat_text_lower for w in item["name_vi"].lower().split())
                    
                    faculty_majors.append({
                        "name_vi": item["name_vi"],
                        "name_ko": item["name_ko"],
                        "is_hot": is_hot or major_kw_match
                    })
                    
            if faculty_majors:
                majors_detail.append({
                    "category": cat_key,
                    "faculty_name_vi": cat_info["name_vi"],
                    "faculty_name_ko": cat_info["name_ko"],
                    "tuition_krw": sem_tuition,
                    "majors": faculty_majors[:4] # 3-4 top majors per faculty
                })
                
        u["majors_detail"] = majors_detail

    # Write back to JS
    js_out = f"export const universities = {json.dumps(universities, ensure_ascii=False, indent=2)};\n"
    with open(path, "w", encoding="utf-8") as f:
        f.write(js_out)
    print(f"Successfully updated {path} with structured majors_detail!")


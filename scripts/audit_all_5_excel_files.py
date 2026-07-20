import openpyxl
import json
import re

with open('scratch/unis.json', 'r', encoding='utf-8') as f:
    unis = json.load(f)

def remove_tones(s):
    if not s: return ''
    s = re.sub(r'[àáạảãâầấậẩẫăằắặẳẵ]', 'a', str(s))
    s = re.sub(r'[èéẹẻẽêềếệểễ]', 'e', str(s))
    s = re.sub(r'[ìíịỉĩ]', 'i', str(s))
    s = re.sub(r'[òóọỏõôồốộổỗơờớợởỡ]', 'o', str(s))
    s = re.sub(r'[ùúụủũưừứựửữ]', 'u', str(s))
    s = re.sub(r'[ỳýỵỷỹ]', 'y', str(s))
    s = re.sub(r'[đ]', 'd', str(s))
    return s.lower().strip()

files = [
    ("TOP1%.xlsx", "TOP 1%"),
    ("TO2%.xlsx", "TOP 2%"),
    ("top3.xlsx", "TOP 3%"),
    ("TruongNoTOPIK.xlsx", "Nợ TOPIK"),
    ("danhsachtruonghanche.xlsx", "Hạn chế Visa")
]

for filename, label in files:
    wb = openpyxl.load_workbook(filename, data_only=True)
    print(f"\n========================================================")
    print(f"📊 KIỂM TRA CHI TIẾT TỆP: {filename} ({label})")
    print(f"========================================================")
    
    all_rows = []
    for sname in wb.sheetnames:
        sheet = wb[sname]
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if not row: continue
            vals = [str(c).strip() for c in row if c is not None and str(c).strip() != '']
            if len(vals) > 0:
                all_rows.append((sname, r_idx, vals))
    
    # Filter out header/title rows
    school_rows = []
    for sname, r_idx, vals in all_rows:
        # A row is a school row if it contains a school name or index number
        joined = " ".join(vals)
        # Check if first or second cell contains school name or number
        if any(kw in joined for kw in ["Đại học", "Cao đẳng", "Viện", "ĐH", "CĐ", "University", "College", "School"]):
            if not any(h in joined.lower() for h in ["danh sách", "tên trường", "bảng học phí", "stt", "loại trường"]):
                school_rows.append((sname, r_idx, vals))

    print(f"Tổng số dòng trường học tìm thấy: {len(school_rows)}")
    
    matched_in_db = 0
    not_matched = []
    
    for sname, r_idx, vals in school_rows:
        # Find school name text in vals
        s_name = ""
        for v in vals:
            if any(kw in v for kw in ["Đại học", "Cao đẳng", "Viện", "ĐH", "CĐ", "University", "College"]):
                if not any(h in v.lower() for h in ["danh sách", "tên trường", "stt"]):
                    s_name = v
                    break
        if not s_name:
            s_name = vals[0] if len(vals) > 0 else ""
        
        # Match with DB
        norm_s = remove_tones(s_name)
        matched = None
        for u in unis:
            norm_vi = remove_tones(u['name_vi'])
            norm_en = remove_tones(u.get('name_en', ''))
            norm_ko = remove_tones(u.get('name_ko', ''))
            if (norm_vi and (norm_vi in norm_s or norm_s in norm_vi)) or (norm_en and norm_en in norm_s) or (norm_ko and norm_ko in norm_s):
                matched = u
                break
        
        if matched:
            matched_in_db += 1
        else:
            not_matched.append(s_name)
            
    print(f"✅ Đã khớp trong DB: {matched_in_db} / {len(school_rows)} trường")
    if not_matched:
        print(f"❌ Chưa khớp trong DB ({len(not_matched)} trường):")
        for nm in not_matched[:15]:
            print(f"   - {nm}")
        if len(not_matched) > 15:
            print(f"   ... và {len(not_matched)-15} trường khác")

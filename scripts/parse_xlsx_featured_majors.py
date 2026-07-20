import openpyxl
import json
import re

xlsx_files = [
    "TOP1%.xlsx",
    "TO2%.xlsx",
    "top3.xlsx"
]

extracted_majors_by_school = {}

for filename in xlsx_files:
    filepath = f"d:/EASS/{filename}"
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb.active
    
    # Find column indices for school name and majors
    name_col = None
    major_col = None
    invoice_col = None
    
    # Check headers in first 4 rows
    for r in range(1, 5):
        for c in range(1, sheet.max_column + 1):
            val = str(sheet.cell(row=r, column=c).value or '').strip().upper()
            if 'TÊN TRƯỜNG' in val or 'TÊN TRƯỜNG' in val or 'TRƯỜNG' in val:
                if not name_col:
                    name_col = c
            if 'THẾ MẠNH' in val or 'NGÀNH NỔI TRỘI' in val or 'NGÀNH NỔI BẬT' in val:
                if not major_col:
                    major_col = c
            if 'INVOICE' in val:
                if not invoice_col:
                    invoice_col = c
                    
    print(f"File {filename} -> Name Col: {name_col}, Major Col: {major_col}, Invoice Col: {invoice_col}")
    
    if name_col and major_col:
        for r in range(2, sheet.max_row + 1):
            s_name = str(sheet.cell(row=r, column=name_col).value or '').strip()
            s_majors = str(sheet.cell(row=r, column=major_col).value or '').strip()
            s_invoice = str(sheet.cell(row=r, column=invoice_col).value or '').strip() if invoice_col else ''
            
            if s_name and s_name.lower() not in ['stt', 'tên trường', 'trường', 'stt']:
                extracted_majors_by_school[s_name] = {
                    "majors_raw": s_majors,
                    "invoice_raw": s_invoice
                }

print(f"\nExtracted data for {len(extracted_majors_by_school)} schools from XLSX files.")
print("Sample extracted entries:")
for k, v in list(extracted_majors_by_school.items())[:15]:
    print(f"  [{k}] -> Majors: {v['majors_raw']}")

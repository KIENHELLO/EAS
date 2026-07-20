import openpyxl
import json
import re
import subprocess

# Load universities from JS via node export
node_script = """
import { universities } from '../src/data/universities.js';
import fs from 'fs';
fs.writeFileSync('scratch/unis.json', JSON.stringify(universities, null, 2), 'utf-8');
"""
with open('scratch/export_unis.js', 'w', encoding='utf-8') as f:
    f.write(node_script)
subprocess.run(['node', 'scratch/export_unis.js'], check=True)

with open('scratch/unis.json', 'r', encoding='utf-8') as f:
    unis = json.load(f)

def normalize_name(s):
    if not s: return ''
    s = str(s).lower().strip()
    s = re.sub(r'[àáạảãâầấậẩẫăằắặẳẵ]', 'a', s)
    s = re.sub(r'[èéẹẻẽêềếệểễ]', 'e', s)
    s = re.sub(r'[ìíịỉĩ]', 'i', s)
    s = re.sub(r'[òóọỏõôồốộổỗơờớợởỡ]', 'o', s)
    s = re.sub(r'[ùúụủũưừứựửữ]', 'u', s)
    s = re.sub(r'[ỳýỵỷỹ]', 'y', s)
    s = re.sub(r'[đ]', 'd', s)
    s = re.sub(r'[\(\)\-\,\.\_\/]', ' ', s)
    words = [w for w in s.split() if w not in ['dai', 'hoc', 'cao', 'dang', 'vien', 'trung', 'cap', 'school', 'university', 'college', 'top', '1%', '2%', '3%']]
    return " ".join(words)

aliases = {
    "sungkyungwan": "skku", "sungkyunkwan": "skku", "chungang": "cau", "chung ang": "cau",
    "pohang": "postech", "busan": "pusan", "seoultech": "seoultech", "sungshin": "sungshin",
    "sookmyung": "sookmyung", "duksung": "duksung", "ewha": "ewha", "khu": "khu", "kyunghee": "khu",
    "gangwon": "kangwon", "kangwon": "kangwon", "soonchunhyang": "uni_8", "hallym": "uni_9",
    "myongji": "uni_10", "sangmyung": "uni_11", "gachon": "uni_1", "dankook": "uni_2",
    "chosun": "uni_3", "donga": "uni_4", "pukyong": "uni_5", "ulsan": "uni_6", "changwon": "uni_7",
    "kyungil": "kyungil", "dong eui": "dongeui", "dongeui": "dongeui", "yong in": "yongin",
    "yongin": "yongin", "dongduk": "dongduk", "mokpo": "mock_uni_120", "sangji": "sangji",
    "chodang": "chodang", "halla": "halla", "keimyung": "keimyung"
}

def find_best_db_match(excel_name):
    norm_ex = normalize_name(excel_name)
    for alias_k, target_id in aliases.items():
        if alias_k in norm_ex or norm_ex == alias_k:
            for u in unis:
                if u['id'] == target_id: return u
    for u in unis:
        norm_vi = normalize_name(u['name_vi'])
        norm_en = normalize_name(u.get('name_en', ''))
        norm_ko = normalize_name(u.get('name_ko', ''))
        if norm_vi and (norm_vi in norm_ex or norm_ex in norm_vi): return u
        if norm_en and (norm_en in norm_ex or norm_ex in norm_en): return u
        if norm_ko and (norm_ko in norm_ex or norm_ex in norm_ko): return u
    return None

def extract_excel_names(filename):
    wb = openpyxl.load_workbook(filename, data_only=True)
    raw_list = []
    for sname in wb.sheetnames:
        sheet = wb[sname]
        for row in sheet.iter_rows(values_only=True):
            if not row: continue
            row_cells = [str(c).strip() for c in row if c is not None and str(c).strip() != '']
            for cell_text in row_cells:
                if any(kw in cell_text for kw in ["Đại học", "Cao đẳng", "Viện", "ĐH", "CĐ", "University", "College"]):
                    if not any(h in cell_text.lower() for h in ["danh sách", "tên trường", "stt", "bảng", "loại trường"]):
                        if cell_text not in raw_list:
                            raw_list.append(cell_text)
                            break
    return raw_list

top1_excel = extract_excel_names('TOP1%.xlsx')
top2_excel = extract_excel_names('TO2%.xlsx')
top3_excel = extract_excel_names('top3.xlsx')
no_topik_excel = extract_excel_names('TruongNoTOPIK.xlsx')
restricted_excel = extract_excel_names('danhsachtruonghanche.xlsx')

top1_ids = set([find_best_db_match(n)['id'] for n in top1_excel if find_best_db_match(n)])
top2_ids = set([find_best_db_match(n)['id'] for n in top2_excel if find_best_db_match(n)])
top3_ids = set([find_best_db_match(n)['id'] for n in top3_excel if find_best_db_match(n)])
no_topik_ids = set([find_best_db_match(n)['id'] for n in no_topik_excel if find_best_db_match(n)])
restricted_ids = set([find_best_db_match(n)['id'] for n in restricted_excel if find_best_db_match(n)])

print(f"============================================================")
print(f"📊 ĐÃ PHÂN TÍCH VÀ KHỚP ID TẤT CẢ FILE EXCEL")
print(f"   • TOP 1% IDs ({len(top1_ids)}): {sorted(list(top1_ids))}")
print(f"   • TOP 2% IDs ({len(top2_ids)}): {sorted(list(top2_ids))}")
print(f"   • TOP 3% IDs ({len(top3_ids)}): {sorted(list(top3_ids))}")
print(f"   • Nợ TOPIK IDs ({len(no_topik_ids)}): {sorted(list(no_topik_ids))}")
print(f"   • Hạn chế Visa IDs ({len(restricted_ids)}): {sorted(list(restricted_ids))}")
print(f"============================================================\n")

# Now update each school in unis
for u in unis:
    u_id = u['id']
    
    # TOP classification
    if u_id in top1_ids:
        u['top_1_percent'] = True
        u['accept_gdtx'] = 'top1'
    elif u_id in top2_ids:
        u['top_1_percent'] = False
        u['accept_gdtx'] = 'top2'
    elif u_id in top3_ids:
        u['top_1_percent'] = False
        u['accept_gdtx'] = 'top3'
    else:
        u['top_1_percent'] = False
        u['accept_gdtx'] = 'top2' # default for unlisted standard schools
        
    # Master No TOPIK
    if u_id in no_topik_ids:
        u['master_no_topik'] = True
    else:
        u['master_no_topik'] = False
        
    # Restricted School
    if u_id in restricted_ids:
        u['is_restricted_school'] = True
    else:
        u['is_restricted_school'] = False

# Save updated unis back to JS files
js_code = "export const universities = " + json.dumps(unis, indent=2, ensure_ascii=False) + ";\n"

with open('src/data/universities.js', 'w', encoding='utf-8') as f:
    f.write(js_code)

with open('kr-unituition-next/src/data/universities.js', 'w', encoding='utf-8') as f:
    f.write(js_code)

print("✅ Đã cập nhật 100% dữ liệu cờ TOP%, Nợ TOPIK, Hạn chế Visa chuẩn từ Excel vào src/data/universities.js và kr-unituition-next/src/data/universities.js!")

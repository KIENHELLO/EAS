import openpyxl
import os

xlsx_files = [
    "TOP1%.xlsx",
    "TO2%.xlsx",
    "top3.xlsx",
    "TruongNoTOPIK.xlsx",
    "danhsachtruonghanche.xlsx"
]

for filename in xlsx_files:
    filepath = os.path.join("d:/EASS", filename)
    if os.path.exists(filepath):
        wb = openpyxl.load_workbook(filepath, data_only=True)
        print(f"\n==========================================")
        print(f"FILE: {filename}")
        print(f"Sheet names: {wb.sheetnames}")
        sheet = wb.active
        
        # Read header row
        headers = []
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=1, column=col).value
            if val is not None:
                headers.append(f"Col {col}: {val}")
        print("Headers (Row 1):")
        for h in headers[:15]:
            print(f"  {h}")
            
        # Inspect first 3 data rows
        print("Sample rows:")
        for r in range(2, min(5, sheet.max_row + 1)):
            row_vals = [str(sheet.cell(row=r, column=c).value or '') for c in range(1, len(headers)+1)]
            print(f"  Row {r}: {row_vals[:8]}")

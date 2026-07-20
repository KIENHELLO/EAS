import openpyxl
import json
import re

with open('scratch/unis.json', 'r', encoding='utf-8') as f:
    unis = json.load(f)

def normalize_name(s):
    if not s: return ''
    s = str(s).lower().strip()
    s = re.sub(r'[àáạảãâầấậẩẫăằắặẳẵ]', 'a', s)
    s = re.sub(r'[èéẹẻẽêềếệểễ]', 'e', s)
    s = re.sub(r'[ìíịỉĩ]', 'i', s)
    s = re.sub(r'[òóọỏõôồốộổỗơờớợởỡ]', 'o', s)
    s = re.sub(r'[ùúụủũưừứựửữ]', 'u', s)
    s = re.sub(r'[ỳýỵỷỹ]', 'y', s)
    s = re.sub(r'[đ]', 'd', s)
    s = re.sub(r'[\(\)\-\,\.\_\/]', ' ', s)
    # Remove common prefix/suffix words
    words = [w for w in s.split() if w not in ['dai', 'hoc', 'cao', 'dang', 'vien', 'trung', 'cap', 'school', 'university', 'college', 'top', '1%', '2%', '3%']]
    return " ".join(words)

# Known aliases dictionary to map Excel names to DB IDs
aliases = {
    "sungkyungwan": "skku",
    "sungkyunkwan": "skku",
    "chungang": "cau",
    "chung ang": "cau",
    "pohang": "postech",
    "busan": "pusan",
    "gwangju": "gwangju",
    "gwangju catholic": "gwangju_catholic",
    "seoultech": "seoultech",
    "khoa hoc cong nghe quoc gia seoul": "seoultech",
    "sungshin": "sungshin",
    "sookmyung": "sookmyung",
    "duksung": "duksung",
    "ewha": "ewha",
    "khu": "khu",
    "kyunghee": "khu",
    "kyung hee": "khu",
    "gangwon": "kangwon",
    "kangwon": "kangwon",
    "soonchunhyang": "uni_8",
    "hallym": "uni_9",
    "myongji": "uni_10",
    "sangmyung": "uni_11",
    "gachon": "uni_1",
    "dankook": "uni_2",
    "chosun": "uni_3",
    "donga": "uni_4",
    "pukyong": "uni_5",
    "ulsan": "uni_6",
    "changwon": "uni_7",
    "kyungil": "kyungil",
    "dong eui": "dongeui",
    "dongeui": "dongeui",
    "yong in": "yongin",
    "yongin": "yongin",
    "dongduk": "dongduk",
    "mokpo": "mock_uni_120",
    "sangji": "sangji",
    "chodang": "chodang",
    "halla": "halla",
    "keimyung": "keimyung"
}

def find_best_db_match(excel_name):
    norm_ex = normalize_name(excel_name)
    
    # 1. Check alias dictionary
    for alias_k, target_id in aliases.items():
        if alias_k in norm_ex or norm_ex == alias_k:
            for u in unis:
                if u['id'] == target_id:
                    return u
    
    # 2. Check DB names
    for u in unis:
        norm_vi = normalize_name(u['name_vi'])
        norm_en = normalize_name(u.get('name_en', ''))
        norm_ko = normalize_name(u.get('name_ko', ''))
        
        if norm_vi and (norm_vi in norm_ex or norm_ex in norm_vi):
            return u
        if norm_en and (norm_en in norm_ex or norm_ex in norm_en):
            return u
        if norm_ko and (norm_ko in norm_ex or norm_ex in norm_ko):
            return u

    return None

# Perform exact audit for all 5 files
files_info = [
    ("TOP1%.xlsx", "top1", "Trường TOP 1%"),
    ("TO2%.xlsx", "top2", "Trường TOP 2%"),
    ("top3.xlsx", "top3", "Trường TOP 3%"),
    ("TruongNoTOPIK.xlsx", "no_topik", "Thạc sĩ / Nợ TOPIK"),
    ("danhsachtruonghanche.xlsx", "restricted", "Hạn chế cấp Visa")
]

audit_summary = {}

for filename, key, title in files_info:
    wb = openpyxl.load_workbook(filename, data_only=True)
    raw_school_list = []
    
    for sname in wb.sheetnames:
        sheet = wb[sname]
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if not row: continue
            row_cells = [str(c).strip() for c in row if c is not None and str(c).strip() != '']
            if not row_cells: continue
            
            # Find name cell
            for cell_text in row_cells:
                if any(kw in cell_text for kw in ["Đại học", "Cao đẳng", "Viện", "ĐH", "CĐ", "University", "College", "School"]):
                    if not any(h in cell_text.lower() for h in ["danh sách", "tên trường", "stt", "bảng", "loại trường"]):
                        if cell_text not in raw_school_list:
                            raw_school_list.append(cell_text)
                            break

    matched_list = []
    unmatched_list = []
    
    for s in raw_school_list:
        db_match = find_best_db_match(s)
        if db_match:
            matched_list.append((s, db_match))
        else:
            unmatched_list.append(s)

    audit_summary[key] = {
        "filename": filename,
        "title": title,
        "total_excel_count": len(raw_school_list),
        "matched_db_count": len(matched_list),
        "unmatched_count": len(unmatched_list),
        "matched": matched_list,
        "unmatched": unmatched_list
    }

print(f"================================================================")
print(f"       KẾT QUẢ ĐỐI CHIẾU THỐNG KÊ CHI TIẾT TỪ FILE EXCEL       ")
print(f"================================================================\n")

for key, data in audit_summary.items():
    print(f"📂 {data['filename']} ({data['title']}):")
    print(f"   • Số trường đọc được từ Excel: {data['total_excel_count']} trường")
    print(f"   • Khớp với cơ sở dữ liệu (DB): {data['matched_db_count']} trường")
    if data['unmatched_count'] > 0:
        print(f"   • Chưa tìm thấy trong DB: {data['unmatched_count']} trường")
        for um in data['unmatched']:
            print(f"      - {um}")
    print("\n")


import openpyxl
import os

xlsx_files = [
    "TOP1%.xlsx",
    "TO2%.xlsx",
    "top3.xlsx",
    "TruongNoTOPIK.xlsx",
    "danhsachtruonghanche.xlsx"
]

for filename in xlsx_files:
    filepath = os.path.join("d:/EASS", filename)
    if os.path.exists(filepath):
        wb = openpyxl.load_workbook(filepath, data_only=True)
        print(f"\n==========================================")
        print(f"FILE: {filename}")
        sheet = wb.active
        print(f"Max row: {sheet.max_row}, Max col: {sheet.max_column}")
        
        # Look for headers in rows 1 to 5
        for r in range(1, min(6, sheet.max_row + 1)):
            row_vals = [str(sheet.cell(row=r, column=c).value or '').strip().replace('\n', ' ') for c in range(1, sheet.max_column + 1)]
            non_empty = [f"Col {idx+1}: {v}" for idx, v in enumerate(row_vals) if v]
            if non_empty:
                print(f"  Row {r}: {non_empty[:10]}")
